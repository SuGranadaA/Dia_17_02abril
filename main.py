#Importamos la librería para archivos de excel
import openpyxl

#Creamos un archivo de Excel
archivo = openpyxl.Workbook()

# Asignamos la hoja de calculo, un objeto de python
hojacalc = archivo.active

#Escribimos contenidos en diferentes casillas de la hoja
hojacalc['A1'] = "Desayuno"
hojacalc['B1'] = "Media mañana"
hojacalc['C1'] = "Almuerzo"
hojacalc['A2'] = "Huevo"
hojacalc['B2'] = "Jamon"
hojacalc['C2'] = "Carne"
hojacalc['A3'] = 160
hojacalc['B3'] = 100
hojacalc['C3'] = 160

#Guardamos el contenido del archivo
archivo.save("hojauno.xlsx")

#Abrimos el archivo
archivo = openpyxl.load_workbook('hojauno.xlsx')

#Imprimimos informacion del XLSX
print(hojacalc['A1'].value)
print(hojacalc['A2'].value)
print(hojacalc['A3'].value)
print(" ")
print(hojacalc['B1'].value)
print(hojacalc['B2'].value)
print(hojacalc['B3'].value)
print(" ")
print(hojacalc['C1'].value)
print(hojacalc['C2'].value)
print(hojacalc['C3'].value)
print("Imprimimos nuevamente")
print(hojacalc.cell(row = 5, column = 2).value)
print(hojacalc.cell(row = 3, column = 4).value)
print(hojacalc.cell(row = 8, column = 6).value)
print(hojacalc.cell(row = 4, column = 7).value)
print(hojacalc.cell(row = 6, column = 5).value)
print(hojacalc.cell(row = 7, column = 4).value)

muchas = hojacalc['B1':'C3']
for row in muchas:
  for column in row:
    print(column.value,end=".")
print(" ")

muchas = hojacalc['E3':'G7']
for row in muchas:
  for column in row:
    print(column.value,end="-")

#Guardamos el archivo
archivo.save("hojauno.xlsx")

#Abrimos el archivo
archivo = openpyxl.load_workbook('hojauno.xlsx')